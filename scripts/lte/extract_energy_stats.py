#!/usr/bin/env python3
"""
extract_energy_stats.py — SBTS Energy Consumption Extractor
RAN Trial Production pipeline — Energy supplement

Reads Nokia SBTS energy consumption Excel exports (one per RC) and computes
period statistics for use in KPI charts and statistical analysis.

IMPORTANT LIMITATION: Nokia energy exports typically cover the trial window
only (no pre-trial baseline period). Period comparison is therefore
Trial vs Post-Rollback, NOT Trial vs Baseline. This must be flagged in
all reports generated from this data.

Metrics extracted:
  - SBTS Energy consumption in BTS (Wh/day total)       → bts_total
  - SBTS Energy consumption in radio modules (Wh/day)   → radio_mod
  - SBTS Energy consumption in system modules (Wh/day)  → sys_mod
  - RU_AVG_PWR_USAGE (W)                                → ru_avg_pwr
  - RU_ENERGY_CONSUMPTION (Wh)                          → ru_energy

Usage:
    python3 extract_energy_stats.py \
        --rc1-file /path/to/RC3_energy.xlsx \
        --rc2-file /path/to/RC4_energy.xlsx \
        --trial-start 2026-02-06 --trial-end 2026-03-04 \
        --post-rb-start 2026-03-05 --post-rb-end 2026-03-17 \
        --rc1-label RC3 --rc2-label RC4 \
        --trial-id CBXXXXXX --out-dir /tmp/CBXXXXXX

Output:
    Prints Python data arrays (ENERGY_RC1, ENERGY_RC2) ready for pasting
    into build_kpi_charts.py and build_stats_report.py.
    Saves energy_summary.csv to --out-dir.
"""

import argparse
import csv
import os
import statistics
from datetime import datetime

try:
    import openpyxl
except ImportError:
    raise SystemExit("openpyxl required: pip install openpyxl --break-system-packages")


# ─── Column indices (0-based) in the Nokia SBTS energy export ──────────────
# Row 0: header labels  Row 1: sub-labels (MRBTS IDs)  Row 2+: data
COL_DATE       = 0
COL_SYS_MOD   = 2   # SBTS Energy consumption in system modules (Wh/day)
COL_RADIO_MOD  = 3   # SBTS Energy consumption in radio modules (Wh/day)
COL_BTS_TOTAL  = 4   # SBTS Energy consumption in BTS (Wh/day)
COL_RU_AVG     = 5   # RU_AVG_PWR_USAGE (W)
COL_RU_ENERGY  = 6   # RU_ENERGY_CONSUMPTION (Wh)
COL_RU_MAX     = 7   # RU_MAX_PWR_USAGE (W)


def load_energy_file(path: str) -> list[dict]:
    """Load a Nokia SBTS energy export. Returns list of daily records."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    records = []
    for row in rows[2:]:  # skip 2 header rows
        if not row or not row[COL_DATE]:
            continue
        dt = row[COL_DATE]
        if not isinstance(dt, datetime):
            continue
        records.append({
            'date':      dt,
            'bts_total': row[COL_BTS_TOTAL],
            'radio_mod': row[COL_RADIO_MOD],
            'sys_mod':   row[COL_SYS_MOD],
            'ru_avg':    row[COL_RU_AVG],
            'ru_energy': row[COL_RU_ENERGY],
        })
    return sorted(records, key=lambda r: r['date'])


def period_stats(records: list[dict], start: datetime, end: datetime) -> dict:
    """Compute mean/std for each metric over [start, end] inclusive."""
    subset = [r for r in records if start <= r['date'] <= end]
    if not subset:
        return {}

    result = {'n': len(subset), 'dates': [r['date'] for r in subset]}
    for metric in ('bts_total', 'radio_mod', 'sys_mod', 'ru_avg', 'ru_energy'):
        vals = [r[metric] for r in subset if r[metric] is not None]
        if vals:
            result[f'{metric}_mean'] = statistics.mean(vals)
            result[f'{metric}_std']  = statistics.stdev(vals) if len(vals) > 1 else 0.0
            result[f'{metric}_vals'] = vals
            result[f'{metric}_daily'] = [
                {'date': r['date'], 'val': r[metric]} for r in subset if r[metric] is not None
            ]
    return result


def sigma(trial_mean, trial_std, post_mean):
    """Sigma of post-RB relative to trial distribution."""
    if trial_std and trial_std > 0:
        return (post_mean - trial_mean) / trial_std
    return 0.0


def parse_date(s: str) -> datetime:
    return datetime.strptime(s, '%Y-%m-%d')


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument('--rc1-file',      required=True)
    ap.add_argument('--rc2-file',      required=True)
    ap.add_argument('--trial-start',   required=True, help='YYYY-MM-DD')
    ap.add_argument('--trial-end',     required=True, help='YYYY-MM-DD')
    ap.add_argument('--post-rb-start', required=True, help='YYYY-MM-DD')
    ap.add_argument('--post-rb-end',   required=True, help='YYYY-MM-DD')
    ap.add_argument('--rc1-label',     default='RC1')
    ap.add_argument('--rc2-label',     default='RC2')
    ap.add_argument('--trial-id',      default='TRIAL')
    ap.add_argument('--out-dir',       default='/tmp/energy_stats')
    args = ap.parse_args()

    t_start  = parse_date(args.trial_start)
    t_end    = parse_date(args.trial_end)
    p_start  = parse_date(args.post_rb_start)
    p_end    = parse_date(args.post_rb_end)

    os.makedirs(args.out_dir, exist_ok=True)

    results = {}
    for label, fpath in [(args.rc1_label, args.rc1_file), (args.rc2_label, args.rc2_file)]:
        print(f"\nLoading {label}: {fpath}")
        data  = load_energy_file(fpath)
        trial = period_stats(data, t_start, t_end)
        post  = period_stats(data, p_start, p_end)
        print(f"  Data range: {data[0]['date'].date()} → {data[-1]['date'].date()}")
        print(f"  Trial days: {trial.get('n',0)}  Post-RB days: {post.get('n',0)}")

        if not trial or not post:
            print(f"  WARNING: insufficient data for period comparison")
            continue

        results[label] = {'data': data, 'trial': trial, 'post': post}

        print(f"\n  {'Metric':<35} {'Trial mean':>12} {'PostRB mean':>12} {'Δ%':>8} {'σ':>6}")
        print(f"  {'─'*75}")
        for metric, unit in [
            ('bts_total', 'Wh/day'),
            ('radio_mod', 'Wh/day'),
            ('sys_mod',   'Wh/day'),
            ('ru_avg',    'W'),
            ('ru_energy', 'Wh'),
        ]:
            tm = trial.get(f'{metric}_mean')
            ts = trial.get(f'{metric}_std', 0)
            pm = post.get(f'{metric}_mean')
            if tm is None or pm is None:
                continue
            dpct = (pm - tm) / tm * 100
            sig  = sigma(tm, ts, pm)
            print(f"  {metric+' ('+unit+')':<35} {tm:>12.1f} {pm:>12.1f} {dpct:>+8.2f}% {sig:>+6.2f}σ")

    # ── Print Python arrays for downstream scripts ─────────────────────────
    print("\n\n# ═══ PASTE INTO build_kpi_charts.py / build_stats_report.py ═══")
    print("# NOTE: energy data covers trial + post-RB only — no baseline period")
    print("# sigma = (post_rb_mean - trial_mean) / trial_std")
    print("# Positive σ = post-RB higher = more energy after rollback (TC=1 uses more)")
    print()

    for label, res in results.items():
        t = res['trial']
        p = res['post']
        varname = f"ENERGY_{label.replace('-','_')}"
        print(f"{varname} = {{")
        for metric, unit in [('bts_total','Wh/day'), ('radio_mod','Wh/day'),
                              ('sys_mod','Wh/day'), ('ru_avg','W'), ('ru_energy','Wh')]:
            tm = t.get(f'{metric}_mean', 0)
            ts = t.get(f'{metric}_std', 0)
            pm = p.get(f'{metric}_mean', 0)
            dpct = (pm - tm) / tm * 100 if tm else 0
            sig  = sigma(tm, ts, pm)
            print(f"    '{metric}': dict(trial={tm:.1f}, post_rb={pm:.1f},"
                  f" delta_pct={dpct:+.2f}, sigma={sig:+.2f}, unit='{unit}'),")
        print(f"    # daily time series for charts:")
        bts_vals = [(r['date'].strftime('%Y-%m-%d'), r['bts_total'])
                    for r in res['data'] if r['bts_total'] is not None]
        print(f"    'bts_total_daily': {bts_vals},")
        print(f"}}")
        print()

    # ── CSV summary ────────────────────────────────────────────────────────
    csv_path = os.path.join(args.out_dir, f"{args.trial_id}_energy_summary.csv")
    rows = [['RC', 'Metric', 'Unit', 'Trial_mean', 'Trial_std',
             'PostRB_mean', 'Delta_pct', 'Sigma']]
    for label, res in results.items():
        t = res['trial']
        p = res['post']
        for metric, unit in [('bts_total','Wh/day'), ('radio_mod','Wh/day'),
                              ('sys_mod','Wh/day'), ('ru_avg','W'), ('ru_energy','Wh')]:
            tm = t.get(f'{metric}_mean', 0)
            ts = t.get(f'{metric}_std', 0)
            pm = p.get(f'{metric}_mean', 0)
            dpct = (pm - tm) / tm * 100 if tm else 0
            sig  = sigma(tm, ts, pm)
            rows.append([label, metric, unit, f'{tm:.1f}', f'{ts:.1f}',
                         f'{pm:.1f}', f'{dpct:+.2f}', f'{sig:+.2f}'])
    with open(csv_path, 'w', newline='') as f:
        csv.writer(f).writerows(rows)
    print(f"CSV saved: {csv_path}")


if __name__ == '__main__':
    main()
